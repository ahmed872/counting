from datetime import datetime

from PyQt6.QtCore import QDate, QSizeF
from PyQt6.QtGui import QTextDocument, QPageSize
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
from PyQt6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QGroupBox,
    QLabel,
    QComboBox,
    QDateEdit,
    QPushButton,
    QTextBrowser,
    QMessageBox,
    QFileDialog,
)

from logic.accounting import AccountingLogic
from logic.hr import HRLogic

PERIOD_DAILY = "يومي"
PERIOD_WEEKLY = "أسبوعي"
PERIOD_MONTHLY = "شهري"
PERIOD_YEARLY = "سنوي"
PERIOD_CUSTOM = "فترة مخصصة"

CATEGORY_LABELS = {
    'raw_material': 'مواد خام',
    'purchase_expense': 'مصروفات مرتبطة بالمشتريات',
    'operating_expense': 'مصروفات تشغيلية',
}


def money(value):
    return f"{value:,.2f}"


class ReportsModule(QWidget):
    """Ready-made period reports (daily / weekly / monthly / yearly / custom)
    that can be printed or saved as PDF."""

    def __init__(self, db_manager):
        super().__init__()
        self.db = db_manager
        self.accounting = AccountingLogic(db_manager)
        self.hr_logic = HRLogic(db_manager)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        header = QLabel("التقارير")
        header.setStyleSheet("font-size: 22px; font-weight: 800; color: #1f3b57;")
        layout.addWidget(header)

        subtitle = QLabel("اختر نوع التقرير والفترة، ثم اطبعه أو احفظه PDF.")
        subtitle.setStyleSheet("color: #64748b;")
        subtitle.setWordWrap(True)
        layout.addWidget(subtitle)

        controls = QGroupBox("إعداد التقرير")
        controls_layout = QVBoxLayout(controls)
        controls_layout.setSpacing(10)

        row = QHBoxLayout()
        row.setSpacing(10)

        self.period_input = QComboBox()
        self.period_input.addItems(
            [PERIOD_DAILY, PERIOD_WEEKLY, PERIOD_MONTHLY, PERIOD_YEARLY, PERIOD_CUSTOM]
        )
        self.period_input.currentTextChanged.connect(self.on_period_changed)

        self.branch_input = QComboBox()
        self.branch_input.addItem("كل الفروع", None)
        for branch in self.db.fetch_all("SELECT id, name FROM branches ORDER BY id"):
            self.branch_input.addItem(branch["name"], branch["id"])

        self.start_date = QDateEdit(QDate.currentDate())
        self.end_date = QDateEdit(QDate.currentDate())

        row.addWidget(self._label("نوع التقرير"))
        row.addWidget(self.period_input, 1)
        row.addSpacing(12)
        row.addWidget(self._label("الفرع"))
        row.addWidget(self.branch_input, 1)
        controls_layout.addLayout(row)

        row2 = QHBoxLayout()
        row2.setSpacing(10)
        row2.addWidget(self._label("من"))
        row2.addWidget(self.start_date, 1)
        row2.addSpacing(12)
        row2.addWidget(self._label("إلى"))
        row2.addWidget(self.end_date, 1)
        controls_layout.addLayout(row2)

        buttons = QHBoxLayout()
        buttons.setSpacing(10)
        generate_btn = QPushButton("عرض التقرير")
        generate_btn.setMinimumHeight(44)
        generate_btn.clicked.connect(self.generate_report)
        pdf_btn = QPushButton("حفظ PDF")
        pdf_btn.setMinimumHeight(44)
        pdf_btn.clicked.connect(self.save_pdf)
        excel_btn = QPushButton("حفظ Excel")
        excel_btn.setMinimumHeight(44)
        excel_btn.clicked.connect(self.save_excel)
        print_btn = QPushButton("طباعة")
        print_btn.setMinimumHeight(44)
        print_btn.clicked.connect(self.print_report)
        buttons.addWidget(generate_btn, 2)
        buttons.addWidget(pdf_btn, 1)
        buttons.addWidget(excel_btn, 1)
        buttons.addWidget(print_btn, 1)
        controls_layout.addLayout(buttons)

        layout.addWidget(controls)

        self.viewer = QTextBrowser()
        self.viewer.setStyleSheet(
            "QTextBrowser { background:white; border:1px solid #dde3ea; border-radius:12px; padding:10px; }"
        )
        self.viewer.setMinimumHeight(320)
        layout.addWidget(self.viewer, 1)

        self.on_period_changed()

    def _label(self, text):
        label = QLabel(text)
        label.setStyleSheet("font-weight:700; color:#334155;")
        return label

    def on_period_changed(self, _text=None):
        custom = self.period_input.currentText() == PERIOD_CUSTOM
        self.start_date.setEnabled(custom)
        self.end_date.setEnabled(custom)
        if not custom:
            start, end = self.resolve_period()
            self.start_date.setDate(QDate.fromString(start, "yyyy-MM-dd"))
            self.end_date.setDate(QDate.fromString(end, "yyyy-MM-dd"))
        self.generate_report()

    def resolve_period(self):
        period = self.period_input.currentText()
        today = QDate.currentDate()
        if period == PERIOD_DAILY:
            return today.toString("yyyy-MM-dd"), today.toString("yyyy-MM-dd")
        if period == PERIOD_WEEKLY:
            return today.addDays(-6).toString("yyyy-MM-dd"), today.toString("yyyy-MM-dd")
        if period == PERIOD_MONTHLY:
            # Whole calendar month, not month-to-date: on the 1st of the month a
            # month-to-date report would show a single day, which is not what
            # "monthly report" means to the person reading it.
            first = QDate(today.year(), today.month(), 1)
            return first.toString("yyyy-MM-dd"), first.addMonths(1).addDays(-1).toString("yyyy-MM-dd")
        if period == PERIOD_YEARLY:
            first = QDate(today.year(), 1, 1)
            return first.toString("yyyy-MM-dd"), QDate(today.year(), 12, 31).toString("yyyy-MM-dd")
        return (self.start_date.date().toString("yyyy-MM-dd"),
                self.end_date.date().toString("yyyy-MM-dd"))

    def _report_data(self):
        start, end = self.resolve_period()
        branch_id = self.branch_input.currentData()
        branch_name = self.branch_input.currentText()
        data = self.accounting.get_period_report(start, end, branch_id)
        # Everything below is a snapshot of where the business stands right
        # now, not filtered to the chosen period - a supplier balance or a
        # loan does not belong to one date range the way a sale does.
        snapshot = {
            'suppliers': self.accounting.get_all_supplier_balances(),
            'customers': self.accounting.get_all_customer_balances(),
            'loans': self.accounting.get_all_loans_with_balances(),
            'prepaid': self.accounting.get_prepaid_expenses_with_balances(),
            'employees': self.hr_logic.get_active_employees_summary(),
            'accrued_wages': self.accounting.get_account_balance('2200'),
            'balance_sheet': self.accounting.get_balance_sheet(),
        }
        return data, branch_name, snapshot

    def current_report_html(self):
        data, branch_name, snapshot = self._report_data()
        return self.build_html(data, self.period_input.currentText(), branch_name, snapshot)

    def generate_report(self):
        try:
            self.viewer.setHtml(self.current_report_html())
        except Exception as exc:
            QMessageBox.critical(self, "خطأ", str(exc))

    def build_html(self, d, period_label, branch_name, snapshot=None):
        sales = d['sales']
        purchases = d['purchases']
        returns = d['returns']
        snapshot = snapshot or {}

        def row(label, value, bold=False, color=None):
            style = "font-weight:800;" if bold else ""
            if color:
                style += f"color:{color};"
            return (f"<tr><td style='padding:6px 10px;{style}'>{label}</td>"
                    f"<td style='padding:6px 10px;text-align:left;{style}'>{money(value)}</td></tr>")

        # Sales and purchases each get their own day-by-day table rather than
        # one row mixing both together with a derived daily profit on top -
        # a sales invoice and a purchase invoice are two different documents,
        # and reading them off one merged row made it hard to hand either
        # side to whoever only needs that one.
        sales_daily_rows = ""
        for entry in d['daily']:
            if not (entry['cash'] or entry['pos'] or entry['transfer'] or entry['delivery']):
                continue
            sales_daily_rows += (
                "<tr>"
                f"<td style='padding:5px 8px'>{entry['day']}</td>"
                f"<td style='padding:5px 8px;text-align:left'>{money(entry['cash'])}</td>"
                f"<td style='padding:5px 8px;text-align:left'>{money(entry['pos'])}</td>"
                f"<td style='padding:5px 8px;text-align:left'>{money(entry['transfer'])}</td>"
                f"<td style='padding:5px 8px;text-align:left'>{money(entry['delivery'])}</td>"
                f"<td style='padding:5px 8px;text-align:left'>{money(entry['sales_total'])}</td>"
                f"<td style='padding:5px 8px;text-align:left'>{money(entry['output_vat'])}</td>"
                "</tr>"
            )
        if not sales_daily_rows:
            sales_daily_rows = ("<tr><td colspan='7' style='padding:12px;text-align:center;color:#64748b'>"
                                "لا توجد مبيعات في هذه الفترة</td></tr>")

        purchases_daily_rows = ""
        for entry in d['daily']:
            if not entry['purchases']:
                continue
            purchases_daily_rows += (
                "<tr>"
                f"<td style='padding:5px 8px'>{entry['day']}</td>"
                f"<td style='padding:5px 8px;text-align:left'>{money(entry['purchases'])}</td>"
                f"<td style='padding:5px 8px;text-align:left'>{money(entry['input_vat'])}</td>"
                "</tr>"
            )
        if not purchases_daily_rows:
            purchases_daily_rows = ("<tr><td colspan='3' style='padding:12px;text-align:center;color:#64748b'>"
                                    "لا توجد مشتريات في هذه الفترة</td></tr>")

        # Only shown when there is something to show - it is always 0 on a
        # branch-filtered report (a credit sale is never attributed to one
        # branch, see get_credit_sales_summary), and an always-zero row on
        # every single-branch report would just be clutter.
        credit_sales_row = (
            row("مبيعات آجلة للعملاء (شاملة الضريبة)", d['credit_sales'] + d['credit_sales_vat'])
            if d['credit_sales'] else ""
        )

        profit_color = "#16a34a" if d['net_profit'] >= 0 else "#dc2626"

        # Everything from here down is "where the business stands right now"
        # rather than a period total - a supplier's balance, a loan, an
        # employee's salary do not belong to the date range picked above.
        suppliers = [s for s in snapshot.get('suppliers', []) if abs(s['balance']) > 0.01]
        suppliers_rows = "".join(
            f"<tr><td style='padding:5px 8px'>{s['name']}{' (متوقف)' if not s['is_active'] else ''}</td>"
            f"<td style='padding:5px 8px;text-align:left'>{money(s['balance'])}</td></tr>"
            for s in suppliers
        ) or ("<tr><td colspan='2' style='padding:12px;text-align:center;color:#64748b'>"
              "لا يوجد مستحق لأي مورد حالياً</td></tr>")
        suppliers_total = sum(s['balance'] for s in suppliers)

        customers = [c for c in snapshot.get('customers', []) if abs(c['balance']) > 0.01]
        customers_rows = "".join(
            f"<tr><td style='padding:5px 8px'>{c['name']}{' (متوقف)' if not c['is_active'] else ''}</td>"
            f"<td style='padding:5px 8px;text-align:left'>{money(c['balance'])}</td></tr>"
            for c in customers
        ) or ("<tr><td colspan='2' style='padding:12px;text-align:center;color:#64748b'>"
              "لا يوجد مستحق على أي عميل حالياً</td></tr>")
        customers_total = sum(c['balance'] for c in customers)

        loans = [l for l in snapshot.get('loans', []) if abs(l['balance']) > 0.01]
        loans_rows = "".join(
            f"<tr><td style='padding:5px 8px'>{l['lender_name']}</td>"
            f"<td style='padding:5px 8px'>{l['date'] or ''}</td>"
            f"<td style='padding:5px 8px;text-align:left'>{money(l['balance'])}</td></tr>"
            for l in loans
        ) or ("<tr><td colspan='3' style='padding:12px;text-align:center;color:#64748b'>"
              "لا يوجد قروض قائمة حالياً</td></tr>")
        loans_total = sum(l['balance'] for l in loans)

        prepaid = [p for p in snapshot.get('prepaid', []) if p['remaining'] > 0.01]
        prepaid_rows = "".join(
            f"<tr><td style='padding:5px 8px'>{p['description']}</td>"
            f"<td style='padding:5px 8px;text-align:left'>{money(p['amount'])}</td>"
            f"<td style='padding:5px 8px;text-align:left'>{money(p['remaining'])}</td></tr>"
            for p in prepaid
        ) or ("<tr><td colspan='3' style='padding:12px;text-align:center;color:#64748b'>"
              "لا يوجد مصروفات مقدمة متبقية حالياً</td></tr>")
        prepaid_total = sum(p['remaining'] for p in prepaid)

        employees = snapshot.get('employees', [])
        employees_rows = "".join(
            f"<tr><td style='padding:5px 8px'>{e['name']}</td>"
            f"<td style='padding:5px 8px'>{e['job_title']}</td>"
            f"<td style='padding:5px 8px'>{e['branch_name']}</td>"
            f"<td style='padding:5px 8px;text-align:left'>{money(e['gross'])}</td></tr>"
            for e in employees
        ) or ("<tr><td colspan='4' style='padding:12px;text-align:center;color:#64748b'>"
              "لا يوجد موظفون نشطون حالياً</td></tr>")
        employees_total = sum(e['gross'] for e in employees)

        bs = snapshot.get('balance_sheet') or {}
        accrued_wages = snapshot.get('accrued_wages', 0)
        bs_balanced_text = "متوازنة ✓" if bs.get('balanced') else "غير متوازنة — راجع القيود"
        bs_balanced_color = "#16a34a" if bs.get('balanced') else "#dc2626"

        return f"""
        <div dir="rtl" style="font-family:'Segoe UI',Tahoma,sans-serif; color:#1f2937;">
          <h2 style="color:#1f3b57;margin-bottom:2px;">تقرير {period_label}</h2>
          <div style="color:#64748b;margin-bottom:4px;">
            الفترة من {d['start_date']} إلى {d['end_date']} &nbsp;|&nbsp; {branch_name}
          </div>
          <div style="color:#94a3b8;font-size:11px;margin-bottom:14px;">
            تم إنشاء التقرير في {datetime.now().strftime('%Y-%m-%d %H:%M')}
          </div>

          <h3 style="color:#1f3b57;">1) المبيعات (شاملة الضريبة)</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;">
            {row("كاش", sales['Cash']['total'])}
            {row("شبكة (مدى / فيزا)", sales['POS']['total'])}
            {row("تحويل بنكي", sales['Transfer']['total'])}
            {row("شركات التوصيل", sales['Delivery']['total'])}
            {row("إجمالي التحصيل", sales['grand_total'], bold=True)}
            {credit_sales_row}
            {row("مرتجعات مبيعات", returns['sales_returns'])}
            {row("صافي المبيعات (بدون ضريبة)", d['net_sales'], bold=True)}
          </table>

          <h3 style="color:#1f3b57;">2) المشتريات والمصروفات (بدون ضريبة)</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;">
            {row(CATEGORY_LABELS['raw_material'], purchases['raw_material']['net'])}
            {row(CATEGORY_LABELS['purchase_expense'], purchases['purchase_expense']['net'])}
            {row(CATEGORY_LABELS['operating_expense'], purchases['operating_expense']['net'])}
            {row("مرتجعات مشتريات", returns['purchase_returns'])}
            {row("إجمالي المشتريات والمصروفات", purchases['grand_net'] - returns['purchase_returns'], bold=True)}
          </table>

          <h3 style="color:#1f3b57;">3) الأرباح</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;">
            {row("صافي المبيعات", d['net_sales'])}
            {row("تكلفة المبيعات", d['cost_of_sales'])}
            {row("مجمل الربح", d['gross_profit'], bold=True)}
            {row("الرواتب والأجور", d['salaries_expense'])}
            {row("المصروفات التشغيلية", d['operating_expenses'])}
            {row("صافي الربح", d['net_profit'], bold=True, color=profit_color)}
          </table>

          <h3 style="color:#1f3b57;">4) ضريبة القيمة المضافة (15%)</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;">
            {row("ضريبة المبيعات (مخرجات)", d['output_vat'])}
            {row("ضريبة المشتريات (مدخلات)", d['input_vat'])}
            {row("صافي الضريبة المستحقة للهيئة", d['net_vat'], bold=True, color="#e67e22")}
          </table>

          <h3 style="color:#1f3b57;">5) تفصيل المبيعات اليومي</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;font-size:12px;">
            <tr style="background:#1f3b57;color:white;">
              <th style="padding:6px 8px;">التاريخ</th>
              <th style="padding:6px 8px;">كاش</th>
              <th style="padding:6px 8px;">شبكة</th>
              <th style="padding:6px 8px;">تحويل</th>
              <th style="padding:6px 8px;">شركات التوصيل</th>
              <th style="padding:6px 8px;">إجمالي المبيعات</th>
              <th style="padding:6px 8px;">ضريبة المبيعات</th>
            </tr>
            {sales_daily_rows}
          </table>

          <h3 style="color:#1f3b57;">6) تفصيل المشتريات اليومي</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;font-size:12px;">
            <tr style="background:#1f3b57;color:white;">
              <th style="padding:6px 8px;">التاريخ</th>
              <th style="padding:6px 8px;">المشتريات</th>
              <th style="padding:6px 8px;">ضريبة المشتريات</th>
            </tr>
            {purchases_daily_rows}
          </table>

          <div style="color:#94a3b8;font-size:11px;margin:16px 0 4px;">
            الأقسام التالية تعرض الأرصدة الحالية وقت إصدار التقرير، وليست محصورة بالفترة المختارة أعلاه.
          </div>

          <h3 style="color:#1f3b57;">7) الموردون - المستحق عليهم حالياً</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;">
            {suppliers_rows}
            <tr><td style='padding:6px 10px;font-weight:800;'>الإجمالي</td>
                <td style='padding:6px 10px;text-align:left;font-weight:800;'>{money(suppliers_total)}</td></tr>
          </table>

          <h3 style="color:#1f3b57;">8) العملاء - المستحق لنا حالياً</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;">
            {customers_rows}
            <tr><td style='padding:6px 10px;font-weight:800;'>الإجمالي</td>
                <td style='padding:6px 10px;text-align:left;font-weight:800;'>{money(customers_total)}</td></tr>
          </table>

          <h3 style="color:#1f3b57;">9) القروض القائمة</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;font-size:12px;">
            <tr style="background:#1f3b57;color:white;">
              <th style="padding:6px 8px;">المُقرض</th>
              <th style="padding:6px 8px;">تاريخ القرض</th>
              <th style="padding:6px 8px;">المتبقي</th>
            </tr>
            {loans_rows}
            <tr><td colspan='2' style='padding:6px 8px;font-weight:800;'>الإجمالي</td>
                <td style='padding:6px 8px;text-align:left;font-weight:800;'>{money(loans_total)}</td></tr>
          </table>

          <h3 style="color:#1f3b57;">10) المصروفات المقدمة المتبقية</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;font-size:12px;">
            <tr style="background:#1f3b57;color:white;">
              <th style="padding:6px 8px;">البيان</th>
              <th style="padding:6px 8px;">المبلغ الكلي</th>
              <th style="padding:6px 8px;">المتبقي</th>
            </tr>
            {prepaid_rows}
            <tr><td style='padding:6px 8px;font-weight:800;'>الإجمالي</td><td></td>
                <td style='padding:6px 8px;text-align:left;font-weight:800;'>{money(prepaid_total)}</td></tr>
          </table>

          <h3 style="color:#1f3b57;">11) طاقم العمل الحالي</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;font-size:12px;">
            <tr style="background:#1f3b57;color:white;">
              <th style="padding:6px 8px;">الاسم</th>
              <th style="padding:6px 8px;">الوظيفة</th>
              <th style="padding:6px 8px;">الفرع</th>
              <th style="padding:6px 8px;">الراتب الإجمالي الشهري</th>
            </tr>
            {employees_rows}
            <tr><td colspan='3' style='padding:6px 8px;font-weight:800;'>الإجمالي</td>
                <td style='padding:6px 8px;text-align:left;font-weight:800;'>{money(employees_total)}</td></tr>
          </table>

          <h3 style="color:#1f3b57;">12) ملخص الوضع المالي العام</h3>
          <table width="100%" cellspacing="0" style="border:1px solid #dde3ea;">
            {row("إجمالي الأصول", bs.get('assets', 0))}
            {row("إجمالي الالتزامات", bs.get('liabilities', 0))}
            {row("منها: رواتب مستحقة الدفع", accrued_wages)}
            {row("حقوق الملكية (شاملة الأرباح المرحّلة)", bs.get('equity', 0))}
            <tr><td style='padding:6px 10px;font-weight:800;'>حالة الميزانية</td>
                <td style='padding:6px 10px;text-align:left;font-weight:800;color:{bs_balanced_color};'>{bs_balanced_text}</td></tr>
          </table>
        </div>
        """

    def _document(self):
        doc = QTextDocument()
        doc.setHtml(self.current_report_html())
        return doc

    def save_pdf(self):
        default = f"تقرير-{self.resolve_period()[0]}.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "حفظ التقرير PDF", default, "PDF (*.pdf)")
        if not path:
            return
        if not path.lower().endswith(".pdf"):
            path += ".pdf"
        try:
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            # Reported live: "حفظ PDF" saved a blank, empty-looking file -
            # but printing the same report and choosing "Microsoft Print to
            # PDF" from the print dialog worked fine. The difference is a
            # real printer: without one set as the Windows default (common -
            # plenty of machines have none), QPrinter has nothing to source
            # a page size from when writing straight to PdfFormat with no
            # printer involved at all, and silently produces a page with
            # degenerate geometry - nothing renders, but the file is not
            # empty either, so this never showed up as an error. An explicit
            # size removes the dependency on there being a real default
            # printer. docs/build_manual.py already does exactly this for
            # the shipped manual, for the same reason.
            printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
            printer.setOutputFileName(path)
            document = self._document()
            document.setPageSize(QSizeF(printer.pageRect(QPrinter.Unit.Point).size()))
            document.print(printer)
            QMessageBox.information(self, "تم", f"تم حفظ التقرير في:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "خطأ", str(exc))

    def print_report(self):
        try:
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            dialog = QPrintDialog(printer, self)
            if dialog.exec() == QPrintDialog.DialogCode.Accepted:
                self._document().print(printer)
        except Exception as exc:
            QMessageBox.critical(self, "خطأ", str(exc))

    def build_workbook(self):
        """The same numbers as the printed report, as real spreadsheet
        rows instead of a page of text - so they can be sorted, filtered,
        or dropped into whatever else the owner already keeps in Excel,
        without retyping a single figure."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        d, branch_name, snapshot = self._report_data()
        bold = Font(bold=True)
        header_font = Font(bold=True, color="FFFFFF")

        wb = Workbook()
        wb.remove(wb.active)

        def sheet(title):
            ws = wb.create_sheet(title)
            ws.sheet_view.rightToLeft = True
            ws.column_dimensions['A'].width = 32
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            return ws

        def header_row(ws, *labels):
            ws.append(labels)
            for cell in ws[ws.max_row]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        summary = sheet("ملخص")
        summary.append([f"تقرير {self.period_input.currentText()}"])
        summary["A1"].font = Font(bold=True, size=14)
        summary.append([f"الفترة من {d['start_date']} إلى {d['end_date']}", branch_name])
        summary.append([])
        for section, rows in (
            ("المبيعات (شاملة الضريبة)", [
                ("كاش", d['sales']['Cash']['total']),
                ("شبكة (مدى / فيزا)", d['sales']['POS']['total']),
                ("تحويل بنكي", d['sales']['Transfer']['total']),
                ("شركات التوصيل", d['sales']['Delivery']['total']),
                ("إجمالي التحصيل", d['sales']['grand_total']),
                ("مرتجعات مبيعات", d['returns']['sales_returns']),
                ("صافي المبيعات (بدون ضريبة)", d['net_sales']),
            ]),
            ("المشتريات والمصروفات (بدون ضريبة)", [
                (CATEGORY_LABELS['raw_material'], d['purchases']['raw_material']['net']),
                (CATEGORY_LABELS['purchase_expense'], d['purchases']['purchase_expense']['net']),
                (CATEGORY_LABELS['operating_expense'], d['purchases']['operating_expense']['net']),
                ("مرتجعات مشتريات", d['returns']['purchase_returns']),
            ]),
            ("الأرباح", [
                ("مجمل الربح", d['gross_profit']),
                ("الرواتب والأجور", d['salaries_expense']),
                ("المصروفات التشغيلية", d['operating_expenses']),
                ("صافي الربح", d['net_profit']),
            ]),
            ("ضريبة القيمة المضافة", [
                ("ضريبة المبيعات (مخرجات)", d['output_vat']),
                ("ضريبة المشتريات (مدخلات)", d['input_vat']),
                ("صافي الضريبة المستحقة", d['net_vat']),
            ]),
        ):
            summary.append([section])
            summary[f"A{summary.max_row}"].font = bold
            for label, value in rows:
                summary.append([label, round(value, 2)])
            summary.append([])

        daily_sales = sheet("المبيعات اليومية")
        header_row(daily_sales, "التاريخ", "كاش", "شبكة", "تحويل", "شركات التوصيل",
                   "إجمالي المبيعات", "ضريبة المبيعات")
        for entry in d['daily']:
            if entry['cash'] or entry['pos'] or entry['transfer'] or entry['delivery']:
                daily_sales.append([entry['day'], round(entry['cash'], 2), round(entry['pos'], 2),
                                    round(entry['transfer'], 2), round(entry['delivery'], 2),
                                    round(entry['sales_total'], 2), round(entry['output_vat'], 2)])

        daily_purchases = sheet("المشتريات اليومية")
        header_row(daily_purchases, "التاريخ", "المشتريات", "ضريبة المشتريات")
        for entry in d['daily']:
            if entry['purchases']:
                daily_purchases.append([entry['day'], round(entry['purchases'], 2), round(entry['input_vat'], 2)])

        suppliers = sheet("الموردون")
        header_row(suppliers, "المورد", "المستحق عليه")
        for s in snapshot.get('suppliers', []):
            if abs(s['balance']) > 0.01:
                suppliers.append([s['name'] + (" (متوقف)" if not s['is_active'] else ""), round(s['balance'], 2)])

        customers = sheet("العملاء")
        header_row(customers, "العميل", "المستحق له")
        for c in snapshot.get('customers', []):
            if abs(c['balance']) > 0.01:
                customers.append([c['name'] + (" (متوقف)" if not c['is_active'] else ""), round(c['balance'], 2)])

        loans = sheet("القروض")
        header_row(loans, "المُقرض", "تاريخ القرض", "المتبقي")
        for l in snapshot.get('loans', []):
            if abs(l['balance']) > 0.01:
                loans.append([l['lender_name'], l['date'] or '', round(l['balance'], 2)])

        prepaid = sheet("المصروفات المقدمة")
        header_row(prepaid, "البيان", "المبلغ الكلي", "المتبقي")
        for p in snapshot.get('prepaid', []):
            if p['remaining'] > 0.01:
                prepaid.append([p['description'], round(p['amount'], 2), round(p['remaining'], 2)])

        employees = sheet("الموظفون")
        header_row(employees, "الاسم", "الوظيفة", "الفرع", "الراتب الإجمالي")
        for e in snapshot.get('employees', []):
            employees.append([e['name'], e['job_title'], e['branch_name'], round(e['gross'], 2)])

        return wb

    def save_excel(self):
        default = f"تقرير-{self.resolve_period()[0]}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "حفظ التقرير Excel", default, "Excel (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            self.build_workbook().save(path)
            QMessageBox.information(self, "تم", f"تم حفظ التقرير في:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "خطأ", str(exc))

    def refresh_on_show(self):
        selected = self.branch_input.currentData()
        self.branch_input.clear()
        self.branch_input.addItem("كل الفروع", None)
        for branch in self.db.fetch_all("SELECT id, name FROM branches ORDER BY id"):
            self.branch_input.addItem(branch["name"], branch["id"])
        index = self.branch_input.findData(selected)
        if index >= 0:
            self.branch_input.setCurrentIndex(index)
        self.generate_report()
